import openpyxl
from openpyxl import Workbook
path = "C:\\Users\\saura\\OneDrive\\Desktop\\Email\\Schools Senior Saturday.xlsx"
wb = openpyxl.load_workbook(path) 
sheet1 = wb.active
correct_answer=5
wrong_answer=-2
participants=sheet1.cell(row=1,column=3)
j=3
participantList=[]
emailList=[]
while(participants.value!=None):
    p=participants.value.split("-")
    participantList.append(p[0])
    emailList.append(p[1].replace(" ",""))
    j=j+1
    participants=sheet1.cell(row=1,column=j)
scores=[]
k=0
while(k<len(participantList)):
    i=3
    key = sheet1.cell(row = i, column = 2)
    attempt = sheet1.cell(row = i, column = k+3)
    score=0
    while(key.value!=None):
        if(attempt.value==None or attempt.value.replace(" ","")==""):
            i=i+1
            attempt = sheet1.cell(row = i, column = k+3)    
            key = sheet1.cell(row = i, column = 2)
            continue
        elif(attempt.value.replace(" ","")==key.value.replace(" ","")):
            score=score+correct_answer
        else:
            score=score+wrong_answer
        i=i+1
        attempt = sheet1.cell(row = i, column = k+3)    
        key = sheet1.cell(row = i, column = 2)
    scores.append(score)
    k=k+1
for i in range(len(scores)): 
    min_idx = i 
    for j in range(i+1, len(scores)): 
        if scores[min_idx] < scores[j]: 
            min_idx = j         
    scores[i], scores[min_idx] = scores[min_idx], scores[i]
    participantList[i],participantList[min_idx]=participantList[min_idx],participantList[i]
    emailList[i],emailList[min_idx]=emailList[min_idx],emailList[i]
sheet2 = wb.create_sheet(index = 2,title = "Ranking")
k=0
while(k<len(participantList)):
    sheet2.cell(row = k+1, column = 1).value=participantList[k]
    sheet2.cell(row=k+1,column=2).value=emailList[k]
    sheet2.cell(row = k+1, column = 3).value=scores[k]
    k=k+1
wb.save("GradedSaturdaySchoolsSeniorQualifier.xlsx")
    
