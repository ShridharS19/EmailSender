import email, smtplib, ssl#install these libraries using pip on python 3.8 or higher
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from PIL import Image, ImageDraw, ImageFont
import openpyxl#install these libraries using pip on python 3.8 or higher
from openpyxl import Workbook

# There are 3 functions : Mail, Cert, and Excel. Mail creates and sends the mail, Cert creates and saves the certificate, and Excel extracts the names and addresses from the excel sheet provided by the site

def Mail(passw,to):#passw stores the app password(a specific password for the gmail account so it can be accessed by outside apps) to is the recipient email address

    subject = ""
    #subject of the mail
    body = ""
    #body of the mail(use \n for new line)
    sender_email = ""
    #sender email id
    receiver_email = to
    password = passw
    

    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = receiver_email
    message["Subject"] = subject
    message["Bcc"] = receiver_email
    message.attach(MIMEText(body, "plain"))#this line till message.attach(part) are to add an attachment to the mail, comment them out if the attachment is not necessary

    filename = "Certificate.pdf"#name of the attachment file(this file must be in the same directory as EmailSender.py)  
    with open(filename, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())

    encoders.encode_base64(part)
    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {filename}",
    )
    message.attach(part)
    text = message.as_string()
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(sender_email, password)
        server.sendmail(sender_email, receiver_email, text)

def Cert(name):#name stores the name of the person to be put on the certificate
    W, H = (3528,2480)#dimensions of the certificate, may wary with different certificates, check with trial and error
    image = Image.open('participation junior.png-page-001.jpg')#participation junior.png-page-001.jpg is the name of the jpg version of the certificate file(this file must be in the same directory as EmailSender.py)  
    draw = ImageDraw.Draw(image)
    font = ImageFont.truetype('Lora.ttf', size=124)#Lora.ttf is a font file(this file must be in the same directory as EmailSender.py)  
    (x, y) = (480, 460)
    message = name
    color = 'rgb(164, 157, 87)'#colour of text
    w, h = draw.textsize(message,font=font)
    
    draw.text(((W-w)/2,(H-h)/2), message, fill=color, font=font)
    im1 = image.convert('RGB')
    im1.save(r"C:\Users\saura\OneDrive\Desktop\Email\Certificate.pdf")#the directory where the edited certificate will be saved as a pdf(must be same directorty as EmailSender.py)

def Excel():
    path = "C:\\Users\\saura\\OneDrive\\Desktop\\Email\\Test.xlsx"#the directory of the excel file with the list of particpants name and email
    wb_obj = openpyxl.load_workbook(path)
    g="60"
    wb_obj.active=wb_obj['Test']#the name of the sheet you want from the excel workbook
    sheet_obj = wb_obj.active
    i=1
    cell_obj = sheet_obj.cell(row = i, column = 2)    
    while(cell_obj!=None):
        Name=sheet_obj.cell(row = i, column = 1)
        print(str(Name.value))
        Cert(Name.value)
        Rec=cell_obj
        Mail("",(Rec.value).replace(" ",""))#enter app password in first parameter
        i=i+1
        cell_obj = sheet_obj.cell(row = i, column = 2)    
        
    
Excel()
